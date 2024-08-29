## Microsoft Planner Transform
# Last update: 2024-08-06


"""About: Create a tasks summary for the most recent export and compare both existing and new tasks marked as completed during each month, saving the output as a Microsoft Excel file."""


###############
# Initial Setup
###############

# Erase all declared global variables
globals().clear()


# Import packages
from datetime import datetime
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import xlwings as xw


# Settings

## Copy-on-Write (will be enabled by default in version 3.0)
if pd.__version__ >= '1.5.0' and pd.__version__ < '3.0.0':
    pd.options.mode.copy_on_write = True


###########
# Functions
###########


def update_or_create_table(*, df, workbook=None, sheet_name, table_name, engine='openpyxl'):
    if engine == 'openpyxl':
        active_sheet = workbook[sheet_name]
        table_exists = any(table.displayName == table_name for table in active_sheet.tables.values())

        if table_exists:
            existing_table = next(table for table in active_sheet.tables.values() if table.displayName == table_name)
            existing_table.ref = f'A1:{get_column_letter(len(df.columns))}{len(df) + 1}'
        else:
            new_table = Table(displayName=table_name, ref=f'A1:{get_column_letter(len(df.columns))}{len(df) + 1}')
            new_table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
            active_sheet.add_table(table=new_table)

    if engine == 'xlwings':
        # Check if the table already exists
        table_exists = False
        for table in sheet_name.tables:
            if table.name == table_name:
                table_exists = True
                # Update the table range
                table.resize(sheet_name.range(f'A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}'))
                break

        if not table_exists:
            # Create a new table
            sheet_name.tables.add(source=sheet_name.range(f'A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}'), name=table_name, table_style_name='TableStyleMedium2')


def microsoft_planner_importer(
    *,
    input_filepath,
    sheet_name,
    labels_mapping=None,
    labels_not_mapped_remove=False,
):
    # Import dataset
    microsoft_planner_checklists_df = (
        pd.read_excel(
            io=input_filepath,
            sheet_name=sheet_name,
            header=0,
            index_col=None,
            skiprows=0,
            skipfooter=0,
            dtype=None,
            engine='openpyxl',
        )
        # Rename columns
        .rename(columns={'task_notes': 'task_description'})
        # Change dtypes
        # .astype(dtype={'run_date': 'datetime64[ns]'})
        .assign(run_month=lambda row: row['run_date'].dt.strftime(date_format='%Y-%m'))
        .assign(
            id=lambda row: row['task_id'].astype(str) + '_' + row['checklist_id'].fillna(value='', method=None, axis=0).astype(str),
        )
        # Reorder columns
        .filter(
            items=[
                'id',
                'run_date',
                'run_month',
                'labels',
                'task_id',
                'task_name',
                'task_description',
                'task_created_date',
                'task_due_date',
                'task_completed_percent',
                'task_completed_date',
                'assigned_to_ids',
                'checklist_id',
                'checklist_name',
                'checklist_value',
            ],
        )
        # Remove duplicate rows
        .drop_duplicates(
            subset=[
                'id',
                'run_date',
                'run_month',
                'labels',
                'task_id',
                'task_name',
                'task_description',
                'task_created_date',
                'task_due_date',
                'task_completed_percent',
                'task_completed_date',
                'assigned_to_ids',
                'checklist_id',
                'checklist_name',
                'checklist_value',
            ],
            keep='first',
            ignore_index=True,
        )
        .assign(
            labels=lambda row: row['labels'].replace(
                to_replace=r'("):true',
                value=r'\1',
                regex=True,
            ),
        )
        .assign(
            labels=lambda row: row['labels'].replace(
                to_replace=r'^{',
                value=r'[',
                regex=True,
            ),
        )
        .assign(
            labels=lambda row: row['labels'].replace(
                to_replace=r'}$',
                value=r']',
                regex=True,
            ),
        )
    )

    if labels_mapping is not None:
        microsoft_planner_checklists_df = microsoft_planner_checklists_df.assign(
            labels=lambda row: row['labels'].replace(
                to_replace=labels_mapping,
                regex=True,
            ),
        )

    if labels_not_mapped_remove is True:
        microsoft_planner_checklists_df = microsoft_planner_checklists_df.assign(
            labels=lambda row: row['labels'].replace(
                to_replace=r'"category[0-9]{1,2}",?',
                value=r'',
                regex=True,
            ),
        )
        microsoft_planner_checklists_df = microsoft_planner_checklists_df.assign(
            labels=lambda row: row['labels'].replace(
                to_replace=r',(])$',
                value=r'\1',
                regex=True,
            ),
        )

    # Return objects
    return microsoft_planner_checklists_df


def microsoft_planner_transform(
    *,
    input_filepath,
    sheet_name,
    labels_mapping=None,
    labels_not_mapped_remove=False,
    output_path,
    file_name,
    engine='openpyxl',
):
    """Create a tasks summary for the most recent export and compare both existing and new tasks marked as completed during each month, saving the output as a Microsoft Excel file."""
    # Create variables
    execution_start = datetime.now()

    # Import and transform Microsoft Planner
    microsoft_planner_tasks_summary_df = microsoft_planner_importer(
        input_filepath=input_filepath,
        sheet_name=sheet_name,
        labels_mapping=labels_mapping,
        labels_not_mapped_remove=labels_not_mapped_remove,
    )

    # Copy dataset
    microsoft_planner_checklists_df = microsoft_planner_tasks_summary_df

    # Tasks Summary
    microsoft_planner_tasks_summary_df = (
        microsoft_planner_tasks_summary_df
        # Keep last 'run_date' rows
        .query(expr='run_date == run_date.max()')
        # Reorder columns
        .filter(
            items=[
                'run_date',
                'run_month',
                'labels',
                'task_id',
                'task_name',
                'task_description',
                'task_created_date',
                'task_due_date',
                'task_completed_percent',
                'task_completed_date',
            ],
        )
        # Remove duplicate rows
        .drop_duplicates(
            subset=['task_id', 'run_date'],
            keep='first',
            ignore_index=True,
        )
        # Reorder rows
        .sort_values(by=['run_date', 'labels', 'task_id'], ignore_index=True)
    )

    ## Checklists Comparer

    # Keep last 'id' value for each 'run_month' rows
    microsoft_planner_checklists_df = pd.merge(
        left=microsoft_planner_checklists_df,
        right=microsoft_planner_checklists_df.groupby(
            by=['run_month', 'id'],
            level=None,
            as_index=False,
            sort=True,
            dropna=True,
        )['run_date'].max(),
        how='inner',
        on=['id', 'run_date', 'run_month'],
        indicator=False,
    ).sort_values(by=['id', 'run_date', 'checklist_value'], ignore_index=True)

    # Add missing 'checklist_value' for 'task_id' without checklists
    microsoft_planner_checklists_df = microsoft_planner_checklists_df.assign(
        checklist_value=lambda row: ((row['checklist_id'].isna()) & (row['checklist_value'].isna()) & (row['task_completed_date'].notna())) | row['checklist_value'],
    )

    # Create 'previous_value' column
    microsoft_planner_checklists_df['previous_value'] = microsoft_planner_checklists_df.groupby(
        by=['id'],
        level=None,
        as_index=False,
        sort=True,
        dropna=True,
    )['checklist_value'].shift(periods=1)

    # Create 'completed' column
    microsoft_planner_checklists_df['completed'] = False

    microsoft_planner_checklists_df = (
        microsoft_planner_checklists_df.assign(
            completed=lambda row: (
                ((row['checklist_value'].eq(True)) & (row['previous_value'].isna()))  # New 'id' checklists marked as completed
                | ((row['checklist_value'].eq(True)) & (row['previous_value'].eq(False)))  # Existing 'id' marked as completed
            ),
        )
        # Remove columns
        .drop(
            columns=['id', 'assigned_to_ids', 'previous_value'],
            axis=1,
            errors='ignore',
        )
        # Reorder rows
        .sort_values(
            by=['run_date', 'labels', 'task_id', 'checklist_id'],
            ignore_index=True,
        )
    )

    if len(microsoft_planner_checklists_df) > 0 or len(microsoft_planner_tasks_summary_df) > 0:
        if engine == 'openpyxl':
            # Check if the file exists
            if os.path.isfile(os.path.join(output_path, file_name)):
                with pd.ExcelWriter(
                    path=os.path.join(output_path, file_name),
                    date_format='YYYY-MM-DD',
                    datetime_format='YYYY-MM-DD',
                    mode='a',
                    if_sheet_exists='overlay',
                    engine='openpyxl',
                    engine_kwargs=None,
                ) as writer:
                    # Replace data in 'Checklists Comparer' sheet
                    if len(microsoft_planner_checklists_df) > 0:
                        microsoft_planner_checklists_df.to_excel(
                            excel_writer=writer,
                            sheet_name='Checklists Comparer',
                            na_rep='',
                            header=True,
                            index=False,
                            index_label=None,
                            freeze_panes=(1, 0),
                            startrow=0,
                            startcol=0,
                        )

                    # Replace data in 'Tasks Summary' sheet
                    if len(microsoft_planner_tasks_summary_df) > 0:
                        microsoft_planner_tasks_summary_df.to_excel(
                            excel_writer=writer,
                            sheet_name='Tasks Summary',
                            na_rep='',
                            header=True,
                            index=False,
                            index_label=None,
                            freeze_panes=(1, 0),
                            startrow=0,
                            startcol=0,
                        )

                    print('')
                    print(f"'{file_name}' file was updated in the '{output_path}' folder.")

            else:
                with pd.ExcelWriter(
                    path=os.path.join(output_path, file_name),
                    date_format='YYYY-MM-DD',
                    datetime_format='YYYY-MM-DD',
                    mode='w',
                    engine='xlsxwriter',
                    engine_kwargs={'options': {'strings_to_formulas': False}},
                ) as writer:
                    if len(microsoft_planner_checklists_df) > 0:
                        microsoft_planner_checklists_df.to_excel(
                            excel_writer=writer,
                            sheet_name='Checklists Comparer',
                            na_rep='',
                            header=True,
                            index=False,
                            index_label=None,
                            freeze_panes=(1, 0),
                        )
                    if len(microsoft_planner_tasks_summary_df) > 0:
                        microsoft_planner_tasks_summary_df.to_excel(
                            excel_writer=writer,
                            sheet_name='Tasks Summary',
                            na_rep='',
                            header=True,
                            index=False,
                            index_label=None,
                            freeze_panes=(1, 0),
                        )

                print('')
                print(f"'{file_name}' file was created and saved to the '{output_path}' folder.")

            if len(microsoft_planner_checklists_df) > 0:
                # Load the workbook
                workbook = load_workbook(filename=os.path.join(output_path, file_name), read_only=False)
                update_or_create_table(df=microsoft_planner_checklists_df, workbook=workbook, sheet_name='Checklists Comparer', table_name='t_checklists_comparer', engine='openpyxl')

                # Save the workbook back to the file
                workbook.save(filename=os.path.join(output_path, file_name))

        if engine == 'xlwings':
            # Ensure the file is not open in Excel
            try:
                app = xw.App(visible=False)
                wb = app.books.open(os.path.join(output_path, file_name))

                # Replace data in 'Checklists Comparer' sheet
                if len(microsoft_planner_checklists_df) > 0:
                    sheet = wb.sheets['Checklists Comparer']
                    sheet.clear()
                    sheet.range('A1').options(index=False).value = microsoft_planner_checklists_df
                    update_or_create_table(df=microsoft_planner_checklists_df, workbook=None, sheet_name=sheet, table_name='t_checklists_comparer', engine='xlwings')

                # Replace data in 'Tasks Summary' sheet
                if len(microsoft_planner_tasks_summary_df) > 0:
                    sheet = wb.sheets['Tasks Summary']
                    sheet.clear()
                    sheet.range('A1').options(index=False).value = microsoft_planner_tasks_summary_df  # Write new data without index

                # Refresh all data connections
                wb.api.RefreshAll()

                # Save and close the workbook
                wb.save(os.path.join(output_path, file_name))
                wb.close()
                app.quit()

                print(f"'{file_name}' file was updated in the '{output_path}' folder.")

            except Exception as e:
                print(f'An error occurred: {e}')
                if 'app' in locals():
                    app.quit()

    # Execution time
    print('')
    print(f'Execution time: {datetime.now() - execution_start}')


def microsoft_excel_pivot_table_refresh(*, input_filepath, sheet_name, pivot_table_name):
    # Load the workbook
    workbook = load_workbook(filename=input_filepath, read_only=False)
    active_sheet = workbook[sheet_name]

    # Find the pivot table
    pivot = next((pt for pt in active_sheet._pivots if pt.name == pivot_table_name), None)

    if pivot is not None:
        # Set the pivot table to refresh on load
        pivot.cache.refreshOnLoad = True

        # Save the workbook back to the file
        workbook.save(filename=input_filepath)

        print('')
        print(f"Pivot table '{pivot_table_name}' was refreshed.")

    else:
        print('')
        print(f"Pivot table '{pivot_table_name}' was not found.")


#############################
# Microsoft Planner Transform
#############################

# Labels mapping dictionary
labels_mapping = {
    '"category1"': '"Category 1"',
    '"category2"': '"Category 2"',
}

# microsoft_planner_transform(
# input_filepath=os.path.join(os.path.expanduser('~'), 'Documents', 'Microsoft Planner Transform', 'Microsoft Planner Export.xlsx'),
# sheet_name='PlannerExport',
# labels_mapping=labels_mapping,
# labels_not_mapped_remove=True,
# output_path=os.path.join(os.path.expanduser('~'), 'Documents', 'Microsoft Planner Transform'),
# file_name='Microsoft Planner Export Transformed (Old).xlsx',
# engine='openpyxl',
# )


microsoft_planner_transform(
    input_filepath=os.path.join(os.path.expanduser('~'), 'Documents', 'Microsoft Planner Transform', 'Microsoft Planner Export.xlsx'),
    sheet_name='PlannerExport',
    labels_mapping=labels_mapping,
    labels_not_mapped_remove=True,
    output_path=os.path.join(os.path.expanduser('~'), 'Documents', 'Microsoft Planner Transform'),
    file_name='Microsoft Planner Export Transformed.xlsx',
    engine='xlwings',
)

# microsoft_excel_pivot_table_refresh(input_filepath=input_filepath=os.path.join(os.path.expanduser('~'), 'Documents', 'Microsoft Planner Transform', 'Microsoft Planner Export Transformed.xlsx'), sheet_name='Report', pivot_table_name='report')
